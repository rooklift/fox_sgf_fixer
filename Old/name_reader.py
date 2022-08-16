import openpyxl

# For decoding the spreadsheet found here:
# https://www.lifein19x19.com/viewtopic.php?p=240400#p240400

class Account():

	def __init__(self, handle, name):
		self.handle = handle
		self.name = name

	def __lt__(self, other):
		return self.handle < other.handle


workbook = openpyxl.load_workbook("FoxGoUsernames.xlsx")
source = workbook.get_sheet_by_name("fox server")

max_column = source.max_column
max_row = source.max_row

outfile = open("outfile.txt", "w", encoding="utf8")

all_accounts = []

for y in range(2, max_row + 1):

	handle = source.cell(column = 1, row = y)
	name1 = source.cell(column = 2, row = y)
	name2 = source.cell(column = 3, row = y)

	if handle.value is None or name1.value is None or name2.value is None:
		continue

	# The spreadsheet has some double handles in a single cell...

	all_handles = handle.value.split("/")

	for h in all_handles:
		account = Account(h.strip(), name1.value + " " + name2.value)
		all_accounts.append(account)

all_accounts.sort()

for account in all_accounts:
	s = '\t"{}":{}"{}",\n'.format(account.handle, (20 - len(account.handle)) * " ", account.name)
	outfile.write(s)
